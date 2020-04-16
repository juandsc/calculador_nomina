import csv
from workalendar.america import Colombia
from configparser import ConfigParser
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl import Workbook


def read_params():
    print('Leyendo parametros...')
    config = ConfigParser()
    config.read('./config.ini')
    params = {}
    params['recargos_semana'] = config.getfloat('recargos', 'semana')
    params['recargos_dom_fest'] = config.getfloat('recargos', 'dom_fest')
    params['extra_diurna_sem'] = config.getfloat('extra', 'diurna_sem')
    params['extra_noct_sem'] = config.getfloat('extra', 'noct_sem')
    params['extra_diurna_dom_fest'] = config.getfloat(
        'extra', 'diurna_dom_fest'
    )
    params['extra_noct_dom_fest'] = config.getfloat('extra', 'noct_dom_fest')
    params['valor_hora_dom_fest'] = config.getfloat('valor_hora', 'dom_fest')
    print('lectura de parametros hecha!\n')
    return params


def read_date(date):
    round_second = round(date.microsecond / 1000000)
    date += timedelta(seconds=round_second)
    return date


def read_hours_file():
    print('Leyendo archivo de horas...')
    wb = load_workbook(filename='./horas.xlsx')
    ws = wb['horas']

    errors = ''
    hours_data = []
    firs_row = True
    row_num = 2
    for row in ws['A{}:C{}'.format(ws.min_row, ws.max_row)]:
        if firs_row:
            firs_row = False
            continue

        if row[0].value is not None:
            current_date = datetime.now()
            # current_year = current_date.year
            current_year = 2018
            # current_month = current_date.month
            start_on = read_date(row[1].value)
            end_on = read_date(row[2].value)
            if start_on.year != current_year and end_on.year != current_year:
                errors += (
                    'Fila '+str(row_num)+': las fechas de entrada y salida no '
                    'son del actual año\n'
                )
            elif start_on > end_on:
                errors += (
                    'Fila '+str(row_num)+': La fecha de entrada es mayor a la '
                    'de salida\n'
                )
            else:
                date_diff = end_on - start_on
                hours_num = date_diff.total_seconds() / 60 / 60
                if hours_num > 24:
                    errors += (
                        'Fila '+str(row_num)+': La diferencia de la fecha de '
                        'entra y salida es mayor a 24 horas\n'
                    )
                else:
                    hours_data.append({
                        'id': row[0].value, 'start_on': start_on,
                        'end_on': end_on
                    })
        row_num += 1

    if errors == '':
        print('lectura de archivo de horas hecha!\n')
    return hours_data, errors


def read_employee_file():
    print('Leyendo informacion de empleados...')
    wb = load_workbook(filename='./empleados.xlsx')
    ws = wb['empleados']

    employee_data = {}
    firs_row = True
    for row in ws['A{}:C{}'.format(ws.min_row, ws.max_row)]:
        if firs_row:
            firs_row = False
            continue

        if row[0].value is not None:
            employee_data[row[0].value] = {
                'id': row[0].value, 'name': row[1].value,
                'salary_base': float(row[2].value),
                'num_diurnal_week': 0,
                'value_diurnal_week': 0,
                'num_extra_diurnal_week': 0,
                'value_extra_diurnal_week': 0,
                'num_extra_noct_week': 0,
                'value_extra_noct_week': 0,
                'num_week_reacharge': 0,
                'value_week_reacharge': 0,
                'num_diurnal_noweek': 0,
                'value_diurnal_noweek': 0,
                'num_extra_diurnal_noweek': 0,
                'value_extra_diurnal_noweek': 0,
                'num_extra_noct_noweek': 0,
                'value_extra_noct_noweek': 0,
                'num_noweek_reacharge': 0,
                'value_noweek_reacharge': 0,
            }
    print('lectura de informacion de empleados hecha!\n')
    return employee_data


def get_type_hour(date):
    if date.weekday() == 5:
        working_day = True
    else:
        working_day = Colombia().is_working_day(date)
    noct_hour = False

    if date.hour >= 21:
        noct_hour = True

    return working_day, noct_hour


def compute_salary_by_hours(hours_data, employee_data, params):
    print('Calculando nomina...')
    for data in hours_data:
        id = data['id']
        employee = employee_data.get(id)
        value_minute = employee['salary_base'] / 240 / 60
        start_on = data['start_on']
        end_on = data['end_on']
        # print(start_on, end_on)
        num_minutes = 1
        while start_on < end_on:
            is_labor_day, is_noct_hour = get_type_hour(start_on)
            if num_minutes <= 480:
                if is_labor_day:
                    if is_noct_hour:
                        employee_data[id]['num_week_reacharge'] += 1 / 60
                        employee_data[id]['value_week_reacharge'] += (
                            value_minute * params['recargos_semana']
                        )
                    else:
                        employee_data[id]['num_diurnal_week'] += 1 / 60
                        employee_data[id]['value_diurnal_week'] += value_minute
                else:
                    if is_noct_hour:
                        employee_data[id]['num_noweek_reacharge'] += 1 / 60
                        employee_data[id]['value_noweek_reacharge'] += (
                            value_minute * params['recargos_dom_fest']
                        )
                    else:
                        employee_data[id]['num_diurnal_noweek'] += 1 / 60
                        employee_data[id]['value_diurnal_noweek'] += (
                            value_minute * params['valor_hora_dom_fest']
                        )
            else:
                if is_labor_day:
                    if is_noct_hour:
                        employee_data[id]['num_extra_noct_week'] += 1 / 60
                        employee_data[id]['value_extra_noct_week'] += (
                            value_minute * params['extra_noct_sem']
                        )
                    else:
                        employee_data[id]['num_extra_diurnal_week'] += 1 / 60
                        employee_data[id]['value_extra_diurnal_week'] += (
                            value_minute * params['extra_diurna_sem']
                        )
                else:
                    if is_noct_hour:
                        employee_data[id]['num_extra_noct_noweek'] += 1 / 60
                        employee_data[id]['value_extra_noct_noweek'] += (
                            value_minute * params['extra_noct_dom_fest']
                        )
                    else:
                        employee_data[id]['num_extra_diurnal_noweek'] += 1 / 60
                        employee_data[id]['value_extra_diurnal_noweek'] += (
                            value_minute * params['extra_diurna_dom_fest']
                        )
            start_on += timedelta(minutes=1)
            num_minutes += 1
        # print('Num minutes:', num_minutes-1)
        if num_minutes/60 < 1:
            print(end_on)

    descriptions = [
        {'Horas Ordinarias': ['num_diurnal_week', 'value_diurnal_week']},
        {'Horas Extras Diurnas': ['num_extra_diurnal_week', 'value_extra_diurnal_week']},
        {'Horas Extras Nocturna': ['num_extra_noct_week', 'value_extra_noct_week']},
        {'Recargos Nocturnos': ['num_week_reacharge', 'value_week_reacharge']},
        {'Horas Domin y Fest': ['num_diurnal_noweek', 'value_diurnal_noweek']},
        {'Horas Extras Diurnas Domin y Fest': ['num_extra_diurnal_noweek', 'value_extra_diurnal_noweek']},
        {'Horas Extras Nocturnas Domin y Fest': ['num_extra_noct_noweek', 'value_extra_noct_noweek']},
        {'Recargos Nocturnos Domin y Fest': ['num_noweek_reacharge', 'value_noweek_reacharge']}
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'salarios calculados'
    for key, employee in employee_data.items():
        ws.append(['CEDULA', '', key])
        ws.append(['NOMBRE', '', employee['name']])
        ws.append(['Descripción', 'Horas', 'Valor Horas'])
        for description in descriptions:
            for key_desc, values in description.items():
                if key_desc == 'Horas Ordinarias':
                    # print('Ordinarias', values)
                    hours = round(employee[values[0]]+16, 2)
                    hours_value = round(employee[values[1]]+(value_minute*960), 2)
                    row = [key_desc, hours, hours_value]
                    ws.append(row)
                elif employee[values[0]] != 0:
                    hours = round(employee[values[0]], 2)
                    hours_value = round(employee[values[1]], 2)
                    row = [key_desc, hours, hours_value]
                    ws.append(row)
        ws.append(['', '', ''])
    wb.save('./horas_procesadas.xlsx')
    print('Calculo de nomina hecha!')


if __name__ == '__main__':
    params = read_params()
    hours_data, errors = read_hours_file()
    if errors != '':
        with open('./errores.txt', 'w') as txt_file:
            txt_file.write(errors)
        print(
            'Mira tu!!... el archivo de horas tiene errores. Mira el archivo '
            'errores.txt para mas información'
        )
    else:
        employee_data = read_employee_file()
        compute_salary_by_hours(hours_data, employee_data, params)

# este es un comentario modificado
